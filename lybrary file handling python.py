from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
import os
from tkinter.ttk import Combobox
from openpyxl import Workbook
import pathlib
from PIL import ImageTk, Image
import openpyxl

root = Tk()
root.title('Library Management System')
root.geometry('900x1200+210+100')
root.config(bg='#a2c1c0')


file_path = pathlib.Path('Library_data.xlsx')
if file_path.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'Book ID'
    sheet['B1'] = 'Title'
    sheet['C1'] = 'Author'
    sheet['D1'] = 'Genre'
    sheet['E1'] = 'Quantity'
    file.save('Library_data.xlsx')

def Exit():
    root.destroy()

def save():
    B1 = BookID.get()
    T1 = Title.get()
    A1 = Author.get()
    G1 = Genre.get()
    Q1 = Quantity.get()

    if T1 == "" or A1 == "" or G1 == "" or Q1 == "":
        messagebox.showerror("Error", "Incomplete Information!")
    else:
        file = openpyxl.load_workbook('Library_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=B1)
        sheet.cell(column=2, row=sheet.max_row, value=T1)
        sheet.cell(column=3, row=sheet.max_row, value=A1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=Q1)
        file.save(r'Library_data.xlsx')
        messagebox.showinfo("Success", "Book registered successfully!")
    Clear()

def search():
    query = Title.get()
    if query == "":
        messagebox.showerror("Error", "Please enter a title to search.")
    else:
        file = openpyxl.load_workbook('Library_data.xlsx')
        sheet = file.active
        found = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value.lower() == query.lower():
                messagebox.showinfo("Book Found", f"Book ID: {sheet.cell(row=row, column=1).value}\n"
                                                  f"Title: {sheet.cell(row=row, column=2).value}\n"
                                                  f"Author: {sheet.cell(row=row, column=3).value}\n"
                                                  f"Genre: {sheet.cell(row=row, column=4).value}\n"
                                                  f"Quantity: {sheet.cell(row=row, column=5).value}")
                found = True
                break
        if not found:
            messagebox.showinfo("Book Not Found", "The book you are looking for is not in the library.")

def Clear():
    BookID.set('')
    Title.set('')
    Author.set('')
    Genre.set('')
    Quantity.set('')

Label(root, text='Email: stefan.mihajlovic@yahoo.com', width=3, height=3, bg="#EDEDED", anchor="e").pack(side=TOP, fill=X)
Label(root, text='Library Management System', width=10, height=2, bg="#c3fbeb", fg='#fff', font='arial 15').pack(side=TOP, fill=X)

obj = LabelFrame(root, text='Book details:', font=20, bd=2, fg='black', bg='#c3fbeb', width=900, height=250, relief=GROOVE)
obj.place(x=30, y=150)

Label(obj, text='Book ID:', font='arial 11', bg='#fb4e30', fg='black').place(x=30, y=50)
Label(obj, text='Title:', font='arial 11', bg='#fb4e30', fg='black').place(x=30, y=100)
Label(obj, text='Author:', font='arial 11', bg='#fb4e30', fg='black').place(x=30, y=150)
Label(obj, text='Genre:', font='arial 11', bg='#fb4e30', fg='black').place(x=500, y=50)
Label(obj, text='Quantity:', font='arial 11', bg='#fb4e30', fg='black').place(x=500, y=100)

BookID = IntVar()
Title = StringVar()
Author = StringVar()
Genre = StringVar()
Quantity = IntVar()

book_id_entry = Entry(obj, textvariable=BookID, font='arial 15', bd=2, width=20)
book_id_entry.place(x=160, y=50)

title_entry = Entry(obj, textvariable=Title, font='arial 15', bd=2, width=20)
title_entry.place(x=160, y=100)

author_entry = Entry(obj, textvariable=Author, font='arial 15', bd=2, width=20)
author_entry.place(x=160, y=150)

genre_entry = Entry(obj, textvariable=Genre, font='arial 15', bd=2, width=20)
genre_entry.place(x=630, y=50)

quantity_entry = Entry(obj, textvariable=Quantity, font='arial 15', bd=2, width=20)
quantity_entry.place(x=630, y=100)

Button(root, text='Save', width=18, height=2, font='arial 12 bold', bg='lightgreen', command=save).place(x=1000, y=150)
Button(root, text='Exit', width=18, height=2, font='arial 12 bold', bg='yellow', command=Exit).place(x=1000, y=230)
Button(root, text='Search', width=18, height=2, font='arial 12 bold', bg='lightblue', command=search).place(x=1000, y=310)

root.mainloop()