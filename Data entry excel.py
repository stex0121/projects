import tkinter
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import pathlib
from openpyxl import Workbook
import openpyxl



root=Tk()
root.resizable(False,False)
root.title('Data Entry')
root.geometry('600x350')
root.configure(bg='#326273')

file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']=' Name'
    sheet['B1']='Last Name'
    sheet['C1']='Phone Number'
    sheet['D1']='Adress'
    sheet['E1']='State'
    sheet['F1']='Gender'

    file.save('Backened_data.xlsx')


def Clear():
    NameValue.set('')
    LastnameValue.set('')
    PhoneNumberValue.set('')
    AdressValue.set('')


def Submit():

    Name=NameValue.get()
    Lastname=LastnameValue.get()
    PhoneNumber=PhoneNumberValue.get()
    Adress=AdressValue.get()
    gender=Gender_combobox.get()
    state=State_combobox.get()
    print(Name)
    print(Lastname)
    print(PhoneNumber)
    print(Adress)
    print(gender)
    print(state)

    file==openpyxl.load_workbook('Backened_Data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=Name)
    sheet.cell(column=2,row=sheet.max_row,value=Lastname)
    sheet.cell(column=3,row=sheet.max_row,value=PhoneNumber)
    sheet.cell(column=4,row=sheet.max_row,value=Adress)
    sheet.cell(column=5,row=sheet.max_row,value=gender)
    sheet.cell(column=6,row=sheet.max_row,value=state)
    file.save(r'Backened_Data.xlsx')
    messagebox.showinfo('info','detail added!')

    NameValue.set('')
    LastnameValue.set('')
    PhoneNumberValue.set('')
    AdressValue.set('')


def Exit():
    root.destroy()




Label(root,text="please fill out this Entry form",font='arial 13',bg='#326273',fg='#fff').place(x=20,y=10)

Label(root,text='Name',font='arial 15',bg='#326273',fg='#fff').place(x=20,y=45)
Label(root,text='Last name',font='arial 15',bg='#326273',fg='#fff').place(x=20,y=90)
Label(root,text='Phone Number',font='arial 15',bg='#326273',fg='#fff').place(x=20,y=135)
Label(root,text='Adress',font='arial 15',bg='#326273',fg='#fff').place(x=20,y=180)
Label(root,text='State',font='arial 15',bg='#326273',fg='#fff').place(x=20,y=225)
Label(root,text='Gender',font='arial 15',bg='#326273',fg='#fff').place(x=350,y=225)

NameValue=StringVar()
LastnameValue=StringVar()
PhoneNumberValue=StringVar()
AdressValue=StringVar()
StateValue=StringVar()
GenderValue=StringVar()

NameEntry=Entry(root,textvariable=NameValue,width=35,bd=2,font=20)
NameEntry.place(x=180,y=45)
Lastname=Entry(root,textvariable=LastnameValue,width=35,bd=2,font=20)
Lastname.place(x=180,y=90)
PhoneNumber=Entry(root,textvariable=PhoneNumberValue,width=35,bd=2,font=20)
PhoneNumber.place(x=180,y=135)
Adress=Entry(root,textvariable=AdressValue,width=35,bd=2,font=20)
Adress.place(x=180,y=180)



Gender_combobox=Combobox(root,values=['Male','Female'],state='r',width=14)
Gender_combobox.place(x=480,y=230)

State_combobox=Combobox(root,values=['England','USA','Brasil','Canada','Spain','France','Serbia',"Germany",'Croatia','Italy'])
State_combobox.place(x=180,y=230)

Button(root,text='Submit',bg='#326273',fg='white',width=15,height=2,command=Submit).place(x=180,y=280)
Button(root,text='Clear',bg='#326273',fg='white',width=15,height=2,command=Clear).place(x=320,y=280)
Button(root,text='Exit',bg='#326273',fg='white',width=15,height=2,command=Exit).place(x=460,y=280)

root.mainloop()

