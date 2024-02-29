from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
import os
from tkinter.ttk import Combobox
from openpyxl import Workbook
import pathlib
from PIL import ImageTk,Image
import openpyxl




root=Tk()
root.title('Student Registration Project')
root.geometry('900x1200+210+100')
root.config(bg='#808080')

file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active()
    sheet['A1']='Registration No.'
    sheet['B1']='Name'
    sheet['C1']='Class'
    sheet['D1']='Gender'
    sheet['E1']='DOB'
    sheet['F1']='Date of Registration'
    sheet['G1']='Religion'
    sheet['H1']='Skill'
    sheet['I1']='Fathers Name'
    sheet['J1']='Mothers Name'
    sheet['K1']='Fathers Occupation'
    sheet['L1']='Mothers Occupation'
    file.save('Student_data.xlsx')



def Exit():
    root.destroy()

def showingimage():
 global filename
 global img
 filename=filedialog.askopenfilename(initialdir=os.getcwd(),title='Select Image file',filetype=('JPG File''*.jpg',('PNG File','*.png'),('All files','*.txt')))
 img=(Image.open(filename))
 resized_image=img.resize((190,190))
 photo2=ImageTk.PhotoImage(resized_image)
 lbl.config(image=photo2)
 lbl.image=photo2

def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    print(max_row_value)
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set('1')

def save():

    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender

    except:

        messagebox.showerror("error", "Select Gender!")

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skills.get()
    fathername = F_name.get()
    mothername = M_name.get()
    F1 = FO_ocuppation.get()
    M1 = Mo_ocuppation.get()
    
    if N1=="" or C1=="" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
     messagebox.showerror("Error", "Few Data is missing!")
    else:

        file=openpyxl.load_workbook('Student_data.xlsx')

        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)
        file.save(r'Student_data.xlsx')
        try:
            img.save("Student Images/"+str(R1)+".png")
        except:
            messagebox.showinfo("info", "Profile Piture is not available!!!!!")
            messagebox.showinfo("info", "Sucessfully data entered!!!!!")
    Clear() 
    registration_no() 




            

def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skills.set('')
    F_name.set('')
    M_name.set('')
    FO_ocuppation.set('')
    Mo_ocuppation.set('')
    Class.set('Select Class')
    registration_no()
    Savebutton.config(state='normal')
    img1=PhotoImage(file='Images/Image upload.png')
    lbl.config(image=img1)
    lbl.image=img1
    img=''



def selection():
    global gender
    value=radio.get()
    if value==1:
        gender='Male'
        print(gender)

def selection1():
    global gender
    value=radio.get()
    if value==1:
        gender='Female'
        print(gender)
    

Label(root,text='Email: stefan.mihajlovic@yahoo.com',width=3,height=3,bg="#EDEDED",anchor="e").pack(side=TOP,fill=X)
Label(root,text='Student Registration',width=10,height=2,bg="#c36464",fg='#fff',font='arial 15').pack(side=TOP,fill=X)



   

Label(root,text='Registration No:',font='arial 13',fg='#EDEDED',bg='#808080').place(x=30,y=150)
Label(root,text='Date:',font='arial 13',fg='#EDEDED',bg='#808080').place(x=500,y=150)

Registration=IntVar()
Date=StringVar()
reg_entry=Entry(root,textvariable=Registration,width=15,font='arial 10')
reg_entry.place(x=160,y=150)

registration_no()

today=date.today()
d1=today.strftime('%d/%m/%Y')
data_entry=Entry(root,textvariable=Date,width=15,font='arial 10')
data_entry.place(x=550,y=150)

Date.set(d1)

obj=LabelFrame(root,text='Students details :',font=20,bd=2,fg='black', bg='#68ddfa',width=900,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text='Full Name:',font='arial 11',bg='#68ddfa',fg='black').place(x=30,y=50)
Label(obj,text='Date of birth:',font='arial 11',bg='#68ddfa',fg='black').place(x=30,y=100)
Label(obj,text='Gender:',font='arial 11',bg='#68ddfa',fg='black').place(x=30,y=150)
Label(obj,text='Class:',font='arial 11',bg='#68ddfa',fg='black').place(x=500,y=50)
Label(obj,text='Religion:',font='arial 11',bg='#68ddfa',fg='black').place(x=500,y=100)
Label(obj,text='Skills:',font='arial 11',bg='#68ddfa',fg='black').place(x=500,y=150)

obj2=LabelFrame(root,text='Parents details :',font=20,bd=2,fg='black', bg='#68ddfa',width=900,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text='Fathers Name:',font='arial 11',bg='#68ddfa',fg='black').place(x=30,y=50)
Label(obj2,text='Occupation:',font='arial 11',bg='#68ddfa',fg='black').place(x=30,y=100)

F_name=StringVar()
f_entry=Entry(obj2,textvariable=F_name,font='arial 15',bd=2 ,width=20 )
f_entry.place(x=160,y=50)

FO_ocuppation=StringVar()
fo_entry=Entry(obj2,textvariable=FO_ocuppation,font='arial 15',bd=2 ,width=20 )
fo_entry.place(x=160,y=100)

Label(obj2,text='Mothers Name:',font='arial 11',bg='#68ddfa',fg='black').place(x=530,y=50)
Label(obj2,text='Occupation:',font='arial 11',bg='#68ddfa',fg='black').place(x=530,y=100)

M_name=StringVar()
m_entry=Entry(obj2,textvariable=M_name,font='arial 15',bd=2 ,width=20 )
m_entry.place(x=660,y=50)

Mo_ocuppation=StringVar()
mo_entry=Entry(obj2,textvariable=Mo_ocuppation,font='arial 15',bd=2 ,width=20 )
mo_entry.place(x=660,y=100)

f=Frame(root,bd=3,bg='black',width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file='upload.png')
lbl=Label(f,bg='black',image=img)
lbl.place(x=0,y=0)

Button(root,text='Upload',width=18,height=2,font='arial 12 bold',bg='lightblue',command=showingimage).place(x=1000,y=370)
Savebutton=Button(root,text='Save',width=18,height=2,font='arial 12 bold',bg='lightgreen',command=save)
Savebutton.place(x=1000,y=450)
Button(root,text='Reset',width=18,height=2,font='arial 12 bold',bg='lightpink',command=Clear).place(x=1000,y=530)
Button(root,text='Exit',width=18,height=2,font='arial 12 bold',bg='yellow',command=Exit).place(x=1000,y=610)

Topimage=PhotoImage(file='topbar.png')
Label(root,image=Topimage).place(x=1000,y=370)

reset=PhotoImage(file='reset.png')
Label(root,image=reset).place(x=1000,y=450)

save=PhotoImage(file='save.png')
Label(root,image=save).place(x=1000,y=530)

exit=PhotoImage(file='exit.png')
Label(root,image=exit).place(x=1000,y=610)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,font='arial 15',bd=2 ,width=20 )
name_entry.place(x=160,y=50)



DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,font='arial 15',bd=2 ,width=20 )
dob_entry.place(x=160,y=100)

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,font='arial 15',bd=2 ,width=20 )
religion_entry.place(x=630,y=100)

radio=IntVar()
R1=Radiobutton(obj,text='Male',variable=radio,value=1,bg='#808080',fg='white',command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text='Female',variable=radio,value=1,bg='#808080',fg='white',command=selection1)
R2.place(x=220,y=150)

Skills=StringVar()
skills_entry=Entry(obj,textvariable=Skills,font='arial 15',bd=2 ,width=20 )
skills_entry.place(x=630,y=150)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8'],font='arial 10',width=17,state='r')
Class.place(x=630,y=50)
Class.set('Select Class')

root.mainloop()

